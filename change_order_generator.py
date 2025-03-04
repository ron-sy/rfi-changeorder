#!/usr/bin/env python3
import os
import json
import pandas as pd
import httpx
from openai import OpenAI
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import PyPDF2
import sys
import time
import logging
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Load environment variables from .env file
load_dotenv()
logger.info("Environment variables loaded")

def extract_text_from_pdf(pdf_path):
    """Extract text from a PDF file with enhanced error handling and text cleaning."""
    logger.info(f"Extracting text from PDF: {pdf_path}")
    try:
        text = ""
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            # Check if PDF is encrypted
            if pdf_reader.is_encrypted:
                logger.warning("PDF is encrypted, attempting to decrypt with empty password")
                try:
                    pdf_reader.decrypt('')  # Try with empty password
                except Exception as e:
                    logger.error(f"Failed to decrypt PDF: {str(e)}")
                    return None
            
            # Extract text from each page
            total_pages = len(pdf_reader.pages)
            logger.info(f"Processing {total_pages} pages")
            
            for page_num in range(total_pages):
                try:
                    page = pdf_reader.pages[page_num]
                    page_text = page.extract_text()
                    
                    # If page text extraction failed, try an alternative method
                    if not page_text or page_text.isspace():
                        logger.warning(f"Page {page_num+1} appears to be empty or contains only images")
                    else:
                        text += page_text + "\n"
                        logger.debug(f"Successfully extracted text from page {page_num+1}")
                except Exception as e:
                    logger.error(f"Failed to extract text from page {page_num+1}: {str(e)}")
        
        # Clean up the extracted text
        text = clean_pdf_text(text)
        
        # Check if we got any meaningful text
        if not text or text.isspace():
            logger.error("No meaningful text extracted from PDF")
            return None
            
        logger.info(f"Successfully extracted {len(text)} characters of text")
        return text
    except FileNotFoundError:
        logger.error(f"PDF file not found: {pdf_path}")
        return None
    except Exception as e:
        logger.error(f"Error extracting text from PDF: {str(e)}", exc_info=True)
        return None

def clean_pdf_text(text):
    """Clean up text extracted from PDF."""
    if not text:
        return ""
        
    # Replace multiple newlines with a single one
    text = '\n'.join(line for line in text.splitlines() if line.strip())
    
    # Replace multiple spaces with a single space
    import re
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'\s+\n', '\n', text)
    text = re.sub(r'\n\s+', '\n', text)
    
    # Remove any strange control characters
    text = ''.join(char for char in text if ord(char) >= 32 or char == '\n')
    
    return text.strip()

def get_user_input():
    """Get the job description from the user, either as text or from a PDF file."""
    print("\n=== Change Order Generator ===")
    print("Choose input method:")
    print("1. Enter job description as text")
    print("2. Provide a PDF file path")
    
    choice = input("Enter your choice (1 or 2): ").strip()
    
    if choice == "1":
        print("\nEnter your job description below (even a general description works).")
        print("Type 'END' on a new line when finished:")
        
        lines = []
        while True:
            line = input()
            if line.strip().upper() == "END":
                break
            lines.append(line)
        
        return "\n".join(lines)
    
    elif choice == "2":
        pdf_path = input("\nEnter the path to your PDF file: ").strip()
        
        # Check if file exists
        if not os.path.exists(pdf_path):
            print(f"Error: File '{pdf_path}' not found.")
            return None
        
        # Check if file is a PDF
        if not pdf_path.lower().endswith('.pdf'):
            print(f"Error: File '{pdf_path}' is not a PDF file.")
            return None
        
        # Extract text from PDF
        text = extract_text_from_pdf(pdf_path)
        if text:
            print("\nSuccessfully extracted text from PDF.")
            # Show a preview of the extracted text
            preview = text[:200] + "..." if len(text) > 200 else text
            print(f"\nPreview of extracted text:\n{preview}")
            
            # Ask if user wants to proceed with this text
            confirm = input("\nProceed with this text? (y/n): ").strip().lower()
            if confirm == 'y':
                return text
            else:
                print("Operation cancelled.")
                return None
        else:
            print("Failed to extract text from PDF.")
            return None
    
    else:
        print("Invalid choice. Please run the program again.")
        return None

def parse_job_description(description):
    """Use OpenAI API to parse the job description into structured data."""
    if not description or not description.strip():
        logger.error("Empty job description provided")
        return None

    try:
        logger.info("Initializing OpenAI client")
        # Initialize httpx client with proper configuration
        http_client = httpx.Client(
            timeout=60.0,
            verify=True  # SSL verification
        )
        
        # Initialize OpenAI client with the http client
        client = OpenAI(
            api_key=os.getenv("OPENAI_API_KEY"),
            http_client=http_client
        )
        
        # Define the system prompt with clearer instructions
        system_prompt = """You are a construction cost estimator. Create a detailed change order from the job description.
Your task is to analyze the job description and create a structured breakdown of costs.

You must return a valid JSON object with this exact structure:
{
    "title": "string",
    "materials": [{"description": "string", "qty": number, "unit": "string", "unit_price": number}],
    "equipment": [{"description": "string", "qty": number, "unit": "string", "duration": number, "duration_unit": "string", "unit_price": number}],
    "labor": [{"description": "string", "trade": "string", "workers": number, "hours_per_day": number, "days": number, "hourly_rate": number}],
    "general_requirements": [{"description": "string", "qty": number, "unit": "string", "unit_price": number}]
}

Important:
1. Return ONLY the JSON object, no other text
2. All numeric values must be numbers, not strings
3. Include realistic quantities and prices
4. Ensure all required fields are present
5. The response must be valid JSON that can be parsed by json.loads()"""

        logger.info(f"Sending request to OpenAI API with description length: {len(description)}")
        
        # Make the API request
        response = client.chat.completions.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": description}
            ],
            temperature=0.7,
            max_tokens=4000
        )
        
        if not response or not response.choices:
            logger.error("Empty response received from OpenAI API")
            return None
            
        content = response.choices[0].message.content
        if not content:
            logger.error("Empty content in API response")
            return None
            
        logger.info("Parsing JSON response from OpenAI API")
        try:
            # Clean the content to ensure it's valid JSON
            content = content.strip()
            if content.startswith('```json'):
                content = content[7:]
            if content.endswith('```'):
                content = content[:-3]
            content = content.strip()
            
            parsed_data = json.loads(content)
            
            # Validate the response has the required structure
            required_keys = ["title", "materials", "equipment", "labor", "general_requirements"]
            missing_keys = [key for key in required_keys if key not in parsed_data]
            if missing_keys:
                logger.error(f"Missing required fields in response: {', '.join(missing_keys)}")
                return None
                
            logger.info("Successfully parsed and validated response")
            return parsed_data
            
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON response: {str(e)}")
            logger.debug(f"Raw response content: {content}")
            return None
            
    except Exception as e:
        logger.error(f"OpenAI API error: {str(e)}", exc_info=True)
        return None

def create_excel_file(data, output_file="change_order.xlsx"):
    """Create an Excel file with the parsed data."""
    logger.info(f"Creating Excel file: {output_file}")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Change Order"
        
        # Define styles
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        total_fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
        grand_total_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
        
        header_font = Font(bold=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        column_widths = [10, 60, 15, 15, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = width
        
        # Add title
        ws.merge_cells('A1:I1')
        ws['A1'] = data.get('title', 'Change Order')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Materials section
        row = 3
        ws['A' + str(row)] = "ITEM"
        ws['B' + str(row)] = "Work Performed By GC (MATERIAL)"
        ws['C' + str(row)] = "QTY"
        ws['D' + str(row)] = "Unit"
        ws['E' + str(row)] = ""
        ws['F' + str(row)] = ""
        ws['G' + str(row)] = ""
        ws['H' + str(row)] = "Unit ($)"
        ws['I' + str(row)] = "Total"
        
        for cell in ws[str(row)]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_alignment
        
        row += 1
        materials_start_row = row
        
        # Add item numbers for materials
        item_num = 1
        for item in data.get('materials', []):
            ws['A' + str(row)] = item_num
            ws['B' + str(row)] = item.get('description', '')
            ws['C' + str(row)] = item.get('qty', '')
            ws['D' + str(row)] = item.get('unit', '')
            ws['H' + str(row)] = f"US$ {item.get('unit_price', 0):.2f}"
            total = item.get('qty', 0) * item.get('unit_price', 0)
            ws['I' + str(row)] = f"US$ {total:.2f}"
            
            for cell in ws[str(row)]:
                cell.border = thin_border
            
            row += 1
            item_num += 1
        
        materials_end_row = row - 1
        
        # Materials Total
        ws['A' + str(row)] = ""
        ws['B' + str(row)] = "Total Direct Cost (Materials)"
        ws['I' + str(row)] = f"US$ {sum(item.get('qty', 0) * item.get('unit_price', 0) for item in data.get('materials', [])):.2f}"
        
        for cell in ws[str(row)]:
            cell.fill = total_fill
            cell.border = thin_border
        
        row += 2
        
        # Equipment section
        ws['A' + str(row)] = "ITEM"
        ws['B' + str(row)] = "Work Performed By GC (EQUIPMENT)"
        ws['C' + str(row)] = "QTY"
        ws['D' + str(row)] = "Unit"
        ws['E' + str(row)] = "Duration"
        ws['F' + str(row)] = "Units"
        ws['G' + str(row)] = ""
        ws['H' + str(row)] = "Unit ($)"
        ws['I' + str(row)] = "Total"
        
        for cell in ws[str(row)]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_alignment
        
        row += 1
        equipment_start_row = row
        
        # Add item numbers for equipment
        item_num = 1
        for item in data.get('equipment', []):
            ws['A' + str(row)] = item_num
            ws['B' + str(row)] = item.get('description', '')
            ws['C' + str(row)] = item.get('qty', '')
            ws['D' + str(row)] = item.get('unit', '')
            ws['E' + str(row)] = item.get('duration', '')
            ws['F' + str(row)] = item.get('duration_unit', '')
            ws['H' + str(row)] = f"US$ {item.get('unit_price', 0):.2f}"
            total = item.get('qty', 0) * item.get('unit_price', 0)
            ws['I' + str(row)] = f"US$ {total:.2f}"
            
            for cell in ws[str(row)]:
                cell.border = thin_border
            
            row += 1
            item_num += 1
        
        equipment_end_row = row - 1
        
        # Equipment Total
        ws['A' + str(row)] = ""
        ws['B' + str(row)] = "Total Direct Cost (Equipment)"
        ws['I' + str(row)] = f"US$ {sum(item.get('qty', 0) * item.get('unit_price', 0) for item in data.get('equipment', [])):.2f}"
        
        for cell in ws[str(row)]:
            cell.fill = total_fill
            cell.border = thin_border
        
        row += 2
        
        # Labor section
        ws['A' + str(row)] = "ITEM"
        ws['B' + str(row)] = "Work Performed By GC (LABOR)"
        ws['C' + str(row)] = "Trade"
        ws['D' + str(row)] = "# of Workers"
        ws['E' + str(row)] = "Hrs/Day"
        ws['F' + str(row)] = "Days"
        ws['G' + str(row)] = "Total Hrs"
        ws['H' + str(row)] = "Hourly Rate"
        ws['I' + str(row)] = "Total"
        
        for cell in ws[str(row)]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_alignment
        
        row += 1
        labor_start_row = row
        
        # Add item numbers for labor
        item_num = 1
        for item in data.get('labor', []):
            ws['A' + str(row)] = item_num
            ws['B' + str(row)] = item.get('description', '')
            ws['C' + str(row)] = item.get('trade', '')
            ws['D' + str(row)] = item.get('workers', '')
            ws['E' + str(row)] = item.get('hours_per_day', '')
            ws['F' + str(row)] = item.get('days', '')
            total_hours = item.get('workers', 0) * item.get('hours_per_day', 0) * item.get('days', 0)
            ws['G' + str(row)] = total_hours
            ws['H' + str(row)] = f"US$ {item.get('hourly_rate', 0):.2f}"
            total = total_hours * item.get('hourly_rate', 0)
            ws['I' + str(row)] = f"US$ {total:.2f}"
            
            for cell in ws[str(row)]:
                cell.border = thin_border
            
            row += 1
            item_num += 1
        
        labor_end_row = row - 1
        
        # Labor Total
        ws['A' + str(row)] = ""
        ws['B' + str(row)] = "Total Direct Cost (Labor)"
        
        labor_total = 0
        for item in data.get('labor', []):
            total_hours = item.get('workers', 0) * item.get('hours_per_day', 0) * item.get('days', 0)
            labor_total += total_hours * item.get('hourly_rate', 0)
        
        ws['I' + str(row)] = f"US$ {labor_total:.2f}"
        
        for cell in ws[str(row)]:
            cell.fill = total_fill
            cell.border = thin_border
        
        row += 2
        
        # Subcontractors section (if any)
        if data.get('subcontractors'):
            ws['A' + str(row)] = "ITEM"
            ws['B' + str(row)] = "Work Performed By ELEC-Sub Contractor"
            ws['C' + str(row)] = "$ Price"
            ws['D' + str(row)] = "QTY"
            ws['E' + str(row)] = "Unit"
            ws['F' + str(row)] = ""
            ws['G' + str(row)] = ""
            ws['H' + str(row)] = ""
            ws['I' + str(row)] = "Total"
            
            for cell in ws[str(row)]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_alignment
            
            row += 1
            subcontractor_start_row = row
            
            # Add item numbers for subcontractors
            item_num = 1
            for item in data.get('subcontractors', []):
                ws['A' + str(row)] = item_num
                ws['B' + str(row)] = item.get('description', '')
                ws['C' + str(row)] = f"US$ {item.get('price', 0):.2f}"
                ws['D' + str(row)] = item.get('qty', '')
                ws['E' + str(row)] = item.get('unit', '')
                total = item.get('qty', 0) * item.get('price', 0)
                ws['I' + str(row)] = f"US$ {total:.2f}"
                
                for cell in ws[str(row)]:
                    cell.border = thin_border
                
                row += 1
                item_num += 1
            
            subcontractor_end_row = row - 1
            
            # Subcontractor Total
            ws['A' + str(row)] = "A"
            ws['B' + str(row)] = "Sub - Contractor's Total Cost"
            
            subcontractor_total = sum(item.get('qty', 0) * item.get('price', 0) for item in data.get('subcontractors', []))
            ws['I' + str(row)] = f"US$ {subcontractor_total:.2f}"
            
            for cell in ws[str(row)]:
                cell.fill = grand_total_fill
                cell.border = thin_border
            
            row += 2
            
            # GC's OH & P on Sub-Contractor's Work
            ws['B' + str(row)] = "GC's OH & P on Sub-Contractor's Work"
            for cell in ws[str(row)]:
                cell.border = thin_border
            
            row += 1
            
            # OH&P tiers
            tier1_limit = 10000
            tier2_limit = 99000
            
            tier1_rate = 0.10  # 10%
            tier2_rate = 0.05  # 5%
            tier3_rate = 0.03  # 3%
            
            tier1_amount = min(subcontractor_total, tier1_limit)
            tier2_amount = min(max(0, subcontractor_total - tier1_limit), tier2_limit - tier1_limit)
            tier3_amount = max(0, subcontractor_total - tier2_limit)
            
            # Tier 1
            ws['B' + str(row)] = f"$ 0 - $ {tier1_limit:,.2f} --({int(tier1_rate*100)}%)"
            ws['C' + str(row)] = f"US$ {tier1_amount:.2f}"
            ws['H' + str(row)] = f"{int(tier1_rate*100)}%"
            ws['I' + str(row)] = f"US$ {tier1_amount * tier1_rate:.2f}"
            for cell in ws[str(row)]:
                cell.border = thin_border
            row += 1
            
            # Tier 2
            ws['B' + str(row)] = f"$ {tier1_limit+1:,.2f} - $ {tier2_limit:,.2f} --({int(tier2_rate*100)}%)"
            ws['C' + str(row)] = f"US$ {tier2_amount:.2f}"
            ws['H' + str(row)] = f"{int(tier2_rate*100)}%"
            ws['I' + str(row)] = f"US$ {tier2_amount * tier2_rate:.2f}"
            for cell in ws[str(row)]:
                cell.border = thin_border
            row += 1
            
            # Tier 3
            ws['B' + str(row)] = f"$ {tier2_limit+1:,.2f} & above --({int(tier3_rate*100)}%)"
            ws['C' + str(row)] = f"US$ {tier3_amount:.2f}"
            ws['H' + str(row)] = f"{int(tier3_rate*100)}%"
            ws['I' + str(row)] = f"US$ {tier3_amount * tier3_rate:.2f}"
            for cell in ws[str(row)]:
                cell.border = thin_border
            row += 1
            
            # OH&P Total
            ws['A' + str(row)] = "B"
            ws['B' + str(row)] = "GC's Overhead & Profit for Subcontractor's Work"
            
            ohp_total = (tier1_amount * tier1_rate) + (tier2_amount * tier2_rate) + (tier3_amount * tier3_rate)
            ws['I' + str(row)] = f"US$ {ohp_total:.2f}"
            
            for cell in ws[str(row)]:
                cell.fill = grand_total_fill
                cell.border = thin_border
            
            row += 2
        
        # General Requirements section (if any)
        if data.get('general_requirements'):
            ws['A' + str(row)] = "ITEM"
            ws['B' + str(row)] = "General Requirements"
            ws['C' + str(row)] = "QTY"
            ws['D' + str(row)] = "Unit"
            ws['E' + str(row)] = ""
            ws['F' + str(row)] = ""
            ws['G' + str(row)] = ""
            ws['H' + str(row)] = "Unit ($)"
            ws['I' + str(row)] = "Total"
            
            for cell in ws[str(row)]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_alignment
            
            row += 1
            general_req_start_row = row
            
            # Add item numbers for general requirements
            item_num = 1
            for item in data.get('general_requirements', []):
                ws['A' + str(row)] = item_num
                ws['B' + str(row)] = item.get('description', '')
                ws['C' + str(row)] = item.get('qty', '')
                ws['D' + str(row)] = item.get('unit', '')
                ws['H' + str(row)] = f"US$ {item.get('unit_price', 0):.2f}"
                total = item.get('qty', 0) * item.get('unit_price', 0)
                ws['I' + str(row)] = f"US$ {total:.2f}"
                
                for cell in ws[str(row)]:
                    cell.border = thin_border
                
                row += 1
                item_num += 1
            
            general_req_end_row = row - 1
            
            # General Requirements Total
            ws['A' + str(row)] = ""
            ws['B' + str(row)] = "Total Direct Cost (General Requirements)"
            ws['I' + str(row)] = f"US$ {sum(item.get('qty', 0) * item.get('unit_price', 0) for item in data.get('general_requirements', [])):.2f}"
            
            for cell in ws[str(row)]:
                cell.fill = total_fill
                cell.border = thin_border
            
            row += 2
        
        # Calculate total direct cost
        materials_total = sum(item.get('qty', 0) * item.get('unit_price', 0) for item in data.get('materials', []))
        equipment_total = sum(item.get('qty', 0) * item.get('unit_price', 0) for item in data.get('equipment', []))
        general_req_total = sum(item.get('qty', 0) * item.get('unit_price', 0) for item in data.get('general_requirements', []))
        
        total_direct_cost = materials_total + equipment_total + labor_total + general_req_total
        
        # Total Direct Cost
        ws['A' + str(row)] = ""
        ws['B' + str(row)] = "Total Direct Cost (Labor + Material + Equipment+ General Requirements)"
        ws['I' + str(row)] = f"US$ {total_direct_cost:.2f}"
        
        for cell in ws[str(row)]:
            cell.fill = grand_total_fill
            cell.border = thin_border
        
        row += 1
        
        # Overhead
        overhead_rate = 0.10  # 10%
        overhead_amount = total_direct_cost * overhead_rate
        
        ws['B' + str(row)] = "Overhead @ 10%"
        ws['H' + str(row)] = "10%"
        ws['I' + str(row)] = f"US$ {overhead_amount:.2f}"
        
        for cell in ws[str(row)]:
            cell.border = thin_border
        
        row += 1
        
        # Profit
        profit_rate = 0.10  # 10%
        profit_amount = total_direct_cost * profit_rate
        
        ws['B' + str(row)] = "Profit @ 10%"
        ws['H' + str(row)] = "10%"
        ws['I' + str(row)] = f"US$ {profit_amount:.2f}"
        
        for cell in ws[str(row)]:
            cell.border = thin_border
        
        row += 1
        
        # Grand Total
        grand_total = total_direct_cost + overhead_amount + profit_amount
        
        if data.get('subcontractors'):
            subcontractor_total = sum(item.get('qty', 0) * item.get('price', 0) for item in data.get('subcontractors', []))
            ohp_total = (min(subcontractor_total, 10000) * 0.10) + (min(max(0, subcontractor_total - 10000), 89000) * 0.05) + (max(0, subcontractor_total - 99000) * 0.03)
            grand_total += subcontractor_total + ohp_total
        
        ws['A' + str(row)] = "C"
        ws['B' + str(row)] = "Grand Total"
        ws['I' + str(row)] = f"US$ {grand_total:.2f}"
        
        for cell in ws[str(row)]:
            cell.fill = grand_total_fill
            cell.border = thin_border
        
        # Add a total row at the very bottom if there are subcontractors
        if data.get('subcontractors'):
            row += 2
            ws['B' + str(row)] = "Total Cost"
            ws['I' + str(row)] = f"US$ {grand_total:.2f}"
            
            for cell in ws[str(row)]:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.border = thin_border
        
        # Save the workbook
        wb.save(output_file)
        logger.info(f"Excel file created successfully at: {os.path.abspath(output_file)}")
        return output_file
    except Exception as e:
        logger.error(f"Failed to create Excel file: {str(e)}", exc_info=True)
        return None

def main():
    """Main function to run the change order generator."""
    logger.info("Starting Change Order Generator")
    
    # Check if a PDF file was provided as a command-line argument
    pdf_path = None
    if len(sys.argv) > 1:
        potential_path = sys.argv[1]
        if os.path.exists(potential_path) and potential_path.lower().endswith('.pdf'):
            pdf_path = potential_path
            logger.info(f"Using provided PDF file: {pdf_path}")
        else:
            logger.warning(f"Provided argument '{potential_path}' is not a valid PDF file path")
    
    print("This program will generate a detailed Excel file based on your job description.")
    print("Even a brief description will be expanded into a comprehensive change order.")
    print("You can enter text directly or provide a PDF file containing the job description.")
    print("Make sure you have set your OpenAI API key in the .env file.")
    
    # Check if API key is set
    if not os.getenv("OPENAI_API_KEY"):
        logger.error("OpenAI API key not found in environment variables")
        print("\nError: OpenAI API key not found!")
        print("Please create a .env file with your API key:")
        print("OPENAI_API_KEY=your_api_key_here")
        return
    
    # Get job description from user or PDF
    job_description = None
    if pdf_path:
        job_description = extract_text_from_pdf(pdf_path)
        if not job_description:
            logger.error(f"Failed to extract text from PDF: {pdf_path}")
            return
    else:
        job_description = get_user_input()
    
    if not job_description:
        logger.error("No job description provided or PDF processing failed")
        return
    
    logger.info("Parsing job description and generating detailed change order")
    parsed_data = parse_job_description(job_description)
    
    if not parsed_data:
        logger.error("Failed to parse job description")
        return
    
    logger.info("Creating Excel file with comprehensive line items")
    output_file = create_excel_file(parsed_data)
    
    if output_file:
        logger.info(f"Change order Excel file created successfully: {output_file}")
    else:
        logger.error("Failed to create Excel file")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.critical("Unexpected error in main execution", exc_info=True)
        print(f"\nAn unexpected error occurred. Please check the log file: {log_filename}") 